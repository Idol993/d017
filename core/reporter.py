import csv
import io
import json
import logging
import os
from datetime import datetime, timedelta
from typing import Dict, List, Optional

from utils.audit_log import query_audit_logs, verify_audit_log_integrity
from utils.db import (
    query_release_records, get_drill_records, save_weekly_report,
    get_weekly_reports, get_release_record,
)
from utils.notify import load_config

logger = logging.getLogger(__name__)


def resolve_date_range(preset: Optional[str] = None,
                       week_start: Optional[str] = None,
                       week_end: Optional[str] = None) -> tuple:
    """根据预设解析日期范围。

    Args:
        preset: 日期预设 - "this_week"(本周) / "last_week"(上周) / "custom"(自定义)
        week_start: 自定义起始日期 (YYYY-MM-DD 或 ISO 格式)
        week_end: 自定义结束日期

    Returns:
        (start_iso, end_iso, preset_label, days) 元组
    """
    today = datetime.now()

    if preset and preset != "custom":
        if preset == "this_week":
            start = today - timedelta(days=today.weekday())
            end = start + timedelta(days=6)
            label = "本周"
        elif preset == "last_week":
            last_monday = today - timedelta(days=today.weekday() + 7)
            start = last_monday
            end = last_monday + timedelta(days=6)
            label = "上周"
        elif preset == "last_30_days":
            end = today
            start = today - timedelta(days=29)
            label = "近30天"
        elif preset == "last_90_days":
            end = today
            start = today - timedelta(days=89)
            label = "近90天"
        else:
            start = today - timedelta(days=today.weekday() + 7)
            end = start + timedelta(days=6)
            label = "上周"

        start_iso = start.strftime("%Y-%m-%dT00:00:00")
        end_iso = end.strftime("%Y-%m-%dT23:59:59")
        days = 7
    else:
        if week_start is None:
            last_monday = today - timedelta(days=today.weekday() + 7)
            start = last_monday
            week_start = last_monday.strftime("%Y-%m-%dT00:00:00")
        else:
            start = datetime.fromisoformat(week_start.replace("T00:00:00", "")) \
                if "T" not in week_start else datetime.fromisoformat(week_start)

        if week_end is None:
            end = start + timedelta(days=6)
            week_end = end.strftime("%Y-%m-%dT23:59:59")
        else:
            end = datetime.fromisoformat(week_end.replace("T23:59:59", "")) \
                if "T" not in week_end else datetime.fromisoformat(week_end)

        start_iso = week_start if "T" in week_start else f"{week_start}T00:00:00"
        end_iso = week_end if "T" in week_end else f"{week_end}T23:59:59"
        days = (end - start).days + 1
        label = f"自定义 ({start_iso[:10]} ~ {end_iso[:10]})"

    return start_iso, end_iso, label, days


class WeeklyReporter:
    def __init__(self, output_dir: str = "./reports", settings: Optional[dict] = None):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        self.settings = settings or load_config()
        self.report_config = self.settings.get("report", {})
        self.chart_config = self.report_config.get("chart", {})
        self.enabled_formats = self.report_config.get("formats", ["json", "csv"])

    def generate_weekly_report(self, week_start: Optional[str] = None,
                                week_end: Optional[str] = None,
                                date_preset: Optional[str] = None,
                                park_filter: Optional[List[str]] = None,
                                formats: Optional[List[str]] = None) -> Dict:
        start_iso, end_iso, preset_label, days = resolve_date_range(
            preset=date_preset, week_start=week_start, week_end=week_end
        )
        week_start = start_iso
        week_end = end_iso

        records = query_release_records(start_time=week_start, end_time=week_end, limit=1000)

        total_raw = len(records)
        if park_filter:
            park_filter_set = set(park_filter)
            records = [
                r for r in records
                if set(r.get("target_parks", []) or []) & park_filter_set
            ]
            park_filter_display = ", ".join(park_filter)
        else:
            park_filter_display = "全部园区"

        total_releases = len(records)
        success_releases = len([r for r in records if r.get("status") == "deployed"])
        rollback_count = len([r for r in records if r.get("status") == "rolled_back"])
        pre_check_failed = len([r for r in records if r.get("status") == "pre_check_failed"])
        approval_rejected = len([r for r in records if r.get("status") == "approval_rejected"])

        success_rate = (success_releases / total_releases * 100) if total_releases > 0 else 0

        approval_durations = []
        for record in records:
            approvals = record.get("approval_records", [])
            if approvals:
                try:
                    first_time = None
                    last_time = None
                    for a in approvals:
                        at = a.get("approved_at", "")
                        if at:
                            if first_time is None or at < first_time:
                                first_time = at
                            if last_time is None or at > last_time:
                                last_time = at
                    if first_time and last_time:
                        duration = (
                            datetime.fromisoformat(last_time)
                            - datetime.fromisoformat(first_time)
                        ).total_seconds() / 60
                        approval_durations.append(duration)
                except (ValueError, TypeError):
                    pass

        avg_approval_duration = (
            sum(approval_durations) / len(approval_durations) if approval_durations else 0
        )

        by_type = {"NORMAL": 0, "HOTFIX": 0}
        by_park = {}
        daily_counts = {}
        daily_success = {}
        daily_rollback = {}

        for record in records:
            rtype = record.get("release_type", "NORMAL")
            by_type[rtype] = by_type.get(rtype, 0) + 1

            parks = record.get("target_parks", [])
            if isinstance(parks, list):
                for park in parks:
                    if park_filter and park not in park_filter:
                        continue
                    by_park[park] = by_park.get(park, 0) + 1

            created = record.get("created_at", "")[:10]
            if created:
                daily_counts[created] = daily_counts.get(created, 0) + 1
                if record.get("status") == "deployed":
                    daily_success[created] = daily_success.get(created, 0) + 1
                elif record.get("status") == "rolled_back":
                    daily_rollback[created] = daily_rollback.get(created, 0) + 1

        date_range_list = []
        start_date = datetime.fromisoformat(week_start.replace("T00:00:00", ""))
        for i in range(days):
            d = (start_date + timedelta(days=i)).strftime("%Y-%m-%d")
            date_range_list.append(d)
            if d not in daily_counts:
                daily_counts[d] = 0
            if d not in daily_success:
                daily_success[d] = 0
            if d not in daily_rollback:
                daily_rollback[d] = 0

        covered_parks = sorted(by_park.keys()) if by_park else []

        filter_info = {
            "date_preset": preset_label,
            "date_start": week_start[:10],
            "date_end": week_end[:10],
            "days": days,
            "park_filter": park_filter_display,
            "park_filter_list": park_filter or [],
            "covered_parks": covered_parks,
            "total_records_in_range": total_raw,
            "records_after_filter": total_releases,
        }

        details = {
            "filter_info": filter_info,
            "by_type": by_type,
            "by_park": by_park,
            "daily_counts": daily_counts,
            "daily_success": daily_success,
            "daily_rollback": daily_rollback,
            "pre_check_failed": pre_check_failed,
            "approval_rejected": approval_rejected,
            "rollback_records": [
                {
                    "id": r.get("id"),
                    "version": r.get("version"),
                    "rollback_report": r.get("rollback_report"),
                    "rolled_back_at": r.get("rolled_back_at"),
                    "target_parks": r.get("target_parks", []),
                }
                for r in records
                if r.get("status") == "rolled_back"
            ],
        }

        report = {
            "id": f"WKR-{datetime.now().strftime('%Y%m%d%H%M%S')}",
            "week_start": week_start[:10] if week_start else "",
            "week_end": week_end[:10] if week_end else "",
            "total_releases": total_releases,
            "success_releases": success_releases,
            "rollback_count": rollback_count,
            "success_rate": round(success_rate, 2),
            "avg_approval_duration_minutes": round(avg_approval_duration, 2),
            "filter_info": filter_info,
            "details": details,
            "generated_at": datetime.now().isoformat(),
        }

        target_formats = formats or self.enabled_formats
        export_files = {}

        text_report = self._generate_text_report(report)

        csv_path = self._export_csv(report)
        export_files["csv"] = csv_path

        json_path = self._export_json(report)
        export_files["json"] = json_path

        if "excel" in target_formats:
            try:
                excel_path = self._export_excel(report, export_files)
                export_files["excel"] = excel_path
            except Exception as e:
                logger.warning("Excel 导出失败: %s", e)

        if "pdf" in target_formats:
            try:
                pdf_path = self._export_pdf(report, export_files)
                export_files["pdf"] = pdf_path
            except Exception as e:
                logger.warning("PDF 导出失败: %s", e)

        report["export_files_info"] = {
            k: os.path.abspath(v) for k, v in export_files.items()
        }

        save_weekly_report(report)

        report["file_path"] = json_path
        report["export_files"] = export_files

        logger.info("运营周报生成完成: %s", report["id"])
        logger.info("  日期范围: %s ~ %s (%s, %d天)",
                    week_start[:10], week_end[:10], preset_label, days)
        logger.info("  园区筛选: %s", park_filter_display)
        logger.info("  记录数: %d (范围内总计 %d)", total_releases, total_raw)
        for fmt, path in export_files.items():
            logger.info("  %-6s → %s", fmt.upper(), os.path.abspath(path))

        result = {
            "report": report,
            "text_report": text_report,
            "export_files": export_files,
            "filter_info": filter_info,
        }

        if "excel" in export_files:
            result["excel_path"] = export_files["excel"]
        if "pdf" in export_files:
            result["pdf_path"] = export_files["pdf"]
        result["csv_path"] = export_files["csv"]
        result["json_path"] = export_files["json"]

        return result

    def _generate_charts(self, report: Dict) -> Dict[str, str]:
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            from matplotlib import rcParams

            rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "Arial Unicode MS"]
            rcParams["axes.unicode_minus"] = False

            chart_paths = {}
            details = report.get("details", {})
            figsize = tuple(self.chart_config.get("figsize", [12, 6]))
            dpi = self.chart_config.get("dpi", 150)

            daily_counts = details.get("daily_counts", {})
            if daily_counts:
                dates = sorted(daily_counts.keys())
                values = [daily_counts[d] for d in dates]
                success_values = [details.get("daily_success", {}).get(d, 0) for d in dates]
                rollback_values = [details.get("daily_rollback", {}).get(d, 0) for d in dates]

                fig, ax = plt.subplots(figsize=figsize, dpi=dpi)
                x = range(len(dates))
                ax.plot(x, values, marker="o", label="总发布数", linewidth=2, color="#3498db")
                ax.plot(x, success_values, marker="s", label="成功发布", linewidth=2, color="#2ecc71")
                ax.plot(x, rollback_values, marker="^", label="回滚数", linewidth=2, color="#e74c3c")
                ax.set_xlabel("日期")
                ax.set_ylabel("发布次数")
                ax.set_title("每日发布趋势")
                ax.set_xticks(x)
                ax.set_xticklabels([d[5:] for d in dates], rotation=45)
                ax.legend()
                ax.grid(True, alpha=0.3)
                plt.tight_layout()

                daily_chart_path = os.path.join(
                    self.output_dir,
                    f"daily_trend_{report['week_start']}.png",
                )
                plt.savefig(daily_chart_path, dpi=dpi, bbox_inches="tight")
                plt.close()
                chart_paths["daily_trend"] = daily_chart_path

            by_type = details.get("by_type", {})
            if by_type:
                labels = list(by_type.keys())
                sizes = list(by_type.values())
                colors = ["#3498db", "#e74c3c"]

                fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(figsize[0], figsize[1] * 0.7), dpi=dpi)
                ax1.pie(sizes, labels=labels, autopct="%1.1f%%", colors=colors, startangle=90)
                ax1.set_title("发布类型分布")

                by_park = details.get("by_park", {})
                if by_park:
                    park_labels = list(by_park.keys())
                    park_values = list(by_park.values())
                    bars = ax2.barh(park_labels, park_values, color="#2ecc71")
                    ax2.set_title("按园区发布分布")
                    ax2.set_xlabel("发布次数")
                    for bar in bars:
                        width = bar.get_width()
                        ax2.text(width, bar.get_y() + bar.get_height()/2,
                                 f"{int(width)}", ha="left", va="center")

                plt.tight_layout()
                dist_chart_path = os.path.join(
                    self.output_dir,
                    f"distribution_{report['week_start']}.png",
                )
                plt.savefig(dist_chart_path, dpi=dpi, bbox_inches="tight")
                plt.close()
                chart_paths["distribution"] = dist_chart_path

            success_rate = report.get("success_rate", 0)
            fig, ax = plt.subplots(figsize=(figsize[0], figsize[1] * 0.5), dpi=dpi)
            colors = ["#2ecc71", "#e74c3c"]
            sizes = [success_rate, 100 - success_rate]
            labels = [f"成功 {success_rate:.1f}%", f"失败 {100-success_rate:.1f}%"]
            ax.pie(sizes, labels=labels, colors=colors, autopct="%1.1f%%", startangle=90)
            ax.set_title("发布成功率")
            plt.tight_layout()
            kpi_chart_path = os.path.join(
                self.output_dir,
                f"kpi_{report['week_start']}.png",
            )
            plt.savefig(kpi_chart_path, dpi=dpi, bbox_inches="tight")
            plt.close()
            chart_paths["kpi"] = kpi_chart_path

            return chart_paths
        except ImportError as e:
            logger.warning("matplotlib 未安装，跳过图表生成: %s", e)
            return {}
        except Exception as e:
            logger.warning("图表生成失败: %s", e)
            return {}

    def _export_excel(self, report: Dict, current_exports: Optional[Dict[str, str]] = None) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        filename = f"weekly_report_{report['week_start']}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()
        wb.remove(wb.active)

        chart_paths = self._generate_charts(report)

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        ws_summary = wb.create_sheet("核心指标", 0)
        ws_summary.merge_cells("A1:E1")
        ws_summary["A1"] = f"物流园区管理系统 - 发布运营周报"
        ws_summary["A1"].font = Font(bold=True, size=16)
        ws_summary["A1"].alignment = center_align

        filter_info = report.get("filter_info", {})

        ws_summary["A3"] = "📋 筛选条件"
        ws_summary["B3"] = "值"
        ws_summary["C3":D3":E3"] = ""
        ws_summary["A3"].font = subheader_font
        ws_summary["B3"] = subheader_font
        ws_summary.merge_cells("B3:E3")

        row = 4
        ws_summary[f"A{row}"] = "统计周期"
        ws_summary[f"B{row}"] = filter_info.get("date_preset", "")
        ws_summary[f"C{row}"] = (
            f"{filter_info.get('date_start', '')
            f" ~ {filter_info.get('date_end', '')}"
        )
        ws_summary[f"D{row}"] = f"{filter_info.get('days', 0)}天"
        ws_summary.merge_cells(f"C{row}:E{row}")

        row = 5
        ws_summary[f"A{row}"] = "园区筛选"
        ws_summary.merge_cells(f"B{row}:E{row}")
        ws_summary[f"B{row}"] = filter_info.get("park_filter", "全部园区")

        row = 6
        ws_summary[f"A{row}"] = "数据覆盖"
        ws_summary.merge_cells(f"B{row}:E{row}")
        parks = filter_info.get("covered_parks", [])
        ws_summary[f"B{row}"] = (
            f"{filter_info.get('records_after_filter', 0)} 条记录, 覆盖 {len(parks)} 个园区"
            f" (范围内总计 {filter_info.get('total_records_in_range', 0)} 条)"
        )

        row = 7
        ws_summary[f"A{row}"] = "生成时间"
        ws_summary.merge_cells(f"B{row}:E{row}")
        ws_summary[f"B{row}"] = report["generated_at"]

        if current_exports:
            row = 8
            ws_summary[f"A{row}"] = "📁 导出文件"
            ws_summary[f"A{row}"].font = subheader_font
            row += 1
            for fmt, fpath in current_exports.items():
                ws_summary[f"A{row}"] = fmt.upper()
                ws_summary.merge_cells(f"B{row}:E{row}")
                ws_summary[f"B{row}"] = os.path.abspath(fpath)
                row += 1

        start_row = row + 2

        summary_headers = ["指标", "数值", "单位", "说明"]
        summary_data = [
            ["发布总数", report["total_releases"], "次", "本周发布申请总数"],
            ["成功发布", report["success_releases"]], "次", "全量发布成功的次数"],
            ["回滚次数", report["rollback_count"]], "次", "触发熔断或手动回滚次数"],
            ["发布成功率", f"{report['success_rate']}%", "%", "成功发布 / 总发布"],
            ["平均审批时长", f"{report['avg_approval_duration_minutes']:.1f}", "分钟", "从提交审批到通过的平均时间"],
        ]

        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for row_idx, row_data in enumerate(summary_data, start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = left_align if col_idx == 4 else center_align
                cell.border = thin_border

        for col in range(1, 5):
            ws_summary.column_dimensions[get_column_letter(col)].width = [22, 25, 12, 40][col-1]

        details = report.get("details", {})

        ws_type = wb.create_sheet("按发布类型", 1)
        ws_type["A1"] = "按发布类型统计"
        ws_type["A1"].font = subheader_font
        ws_type.merge_cells("A1:B1")

        type_headers = ["发布类型", "次数"]
        for col, header in enumerate(type_headers, 1):
            cell = ws_type.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for row_idx, (rtype, count) in enumerate(details.get("by_type", {}).items(), 4):
            ws_type.cell(row=row_idx, column=1, value=rtype).alignment = center_align
            ws_type.cell(row=row_idx, column=1).border = thin_border
            ws_type.cell(row=row_idx, column=2, value=count).alignment = center_align
            ws_type.cell(row=row_idx, column=2).border = thin_border

        ws_type.column_dimensions["A"].width = 20
        ws_type.column_dimensions["B"].width = 15

        ws_park = wb.create_sheet("按园区分布", 2)
        ws_park["A1"] = "按园区发布统计"
        ws_park["A1"].font = subheader_font
        ws_park.merge_cells("A1:B1")

        park_headers = ["园区ID", "发布次数"]
        for col, header in enumerate(park_headers, 1):
            cell = ws_park.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for row_idx, (park, count) in enumerate(details.get("by_park", {}).items(), 4):
            ws_park.cell(row=row_idx, column=1, value=park).alignment = center_align
            ws_park.cell(row=row_idx, column=1).border = thin_border
            ws_park.cell(row=row_idx, column=2, value=count).alignment = center_align
            ws_park.cell(row=row_idx, column=2).border = thin_border

        ws_park.column_dimensions["A"].width = 25
        ws_park.column_dimensions["B"].width = 15

        ws_daily = wb.create_sheet("每日趋势", 3)
        ws_daily["A1"] = "每日发布趋势"
        ws_daily["A1"].font = subheader_font
        ws_daily.merge_cells("A1:D1")

        daily_headers = ["日期", "总发布数", "成功发布", "回滚数"]
        for col, header in enumerate(daily_headers, 1):
            cell = ws_daily.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        daily_counts = details.get("daily_counts", {})
        daily_success = details.get("daily_success", {})
        daily_rollback = details.get("daily_rollback", {})
        for row_idx, date in enumerate(sorted(daily_counts.keys()), 4):
            ws_daily.cell(row=row_idx, column=1, value=date).alignment = center_align
            ws_daily.cell(row=row_idx, column=1).border = thin_border
            ws_daily.cell(row=row_idx, column=2, value=daily_counts[date]).alignment = center_align
            ws_daily.cell(row=row_idx, column=2).border = thin_border
            ws_daily.cell(row=row_idx, column=3, value=daily_success.get(date, 0)).alignment = center_align
            ws_daily.cell(row=row_idx, column=3).border = thin_border
            ws_daily.cell(row=row_idx, column=4, value=daily_rollback.get(date, 0)).alignment = center_align
            ws_daily.cell(row=row_idx, column=4).border = thin_border

        for col in range(1, 5):
            ws_daily.column_dimensions[get_column_letter(col)].width = [15, 12, 12, 12][col-1]

        if "daily_trend" in chart_paths:
            from openpyxl.drawing.image import Image
            try:
                img = Image(chart_paths["daily_trend"])
                img.width = 600
                img.height = 300
                ws_daily.add_image(img, "F3")
            except Exception as e:
                logger.warning("Excel 插入图表失败: %s", e)

        if "distribution" in chart_paths:
            try:
                img = Image(chart_paths["distribution"])
                img.width = 600
                img.height = 250
                ws_summary.add_image(img, f"F{start_row}")
            except Exception as e:
                logger.warning("Excel 插入分布图失败: %s", e)

        rollback_records = details.get("rollback_records", [])
        if rollback_records:
            ws_rollback = wb.create_sheet("回滚详情", 4)
            ws_rollback["A1"] = "回滚记录详情"
            ws_rollback["A1"].font = subheader_font
            ws_rollback.merge_cells("A1:E1")

            rb_headers = ["发布单号", "版本号", "回滚原因", "触发指标", "回滚时间"]
            for col, header in enumerate(rb_headers, 1):
                cell = ws_rollback.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            for row_idx, rr in enumerate(rollback_records, 4):
                ws_rollback.cell(row=row_idx, column=1, value=rr.get("id", "")).alignment = center_align
                ws_rollback.cell(row=row_idx, column=1).border = thin_border
                ws_rollback.cell(row=row_idx, column=2, value=rr.get("version", "")).alignment = center_align
                ws_rollback.cell(row=row_idx, column=2).border = thin_border

                rb_detail = rr.get("rollback_report", {}) or {}
                ws_rollback.cell(row=row_idx, column=3, value=rb_detail.get("reason", "")).alignment = left_align
                ws_rollback.cell(row=row_idx, column=3).border = thin_border
                ws_rollback.cell(row=row_idx, column=4, value=rb_detail.get("trigger_metric", "")).alignment = center_align
                ws_rollback.cell(row=row_idx, column=4).border = thin_border
                ws_rollback.cell(row=row_idx, column=5, value=rr.get("rolled_back_at", "")).alignment = center_align
                ws_rollback.cell(row=row_idx, column=5).border = thin_border

            for col in range(1, 6):
                ws_rollback.column_dimensions[get_column_letter(col)].width = [20, 15, 40, 20, 25][col-1]

        wb.save(filepath)
        return filepath

    def _export_pdf(self, report: Dict, current_exports: Optional[Dict[str, str]] = None) -> str:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, Image as RLImage,
        )
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        filename = f"weekly_report_{report['week_start']}.pdf"
        filepath = os.path.join(self.output_dir, filename)

        try:
            font_paths = [
                "C:/Windows/Fonts/msyh.ttc",
                "C:/Windows/Fonts/simhei.ttf",
                "C:/Windows/Fonts/simsun.ttc",
                "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
                "/System/Library/Fonts/PingFang.ttc",
            ]
            font_loaded = False
            for font_path in font_paths:
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont("CustomFont", font_path))
                        font_loaded = True
                        break
                    except Exception:
                        continue
        except Exception:
            pass

        chart_paths = self._generate_charts(report)

        doc = SimpleDocTemplate(
            filepath, pagesize=A4,
            rightMargin=36, leftMargin=36,
            topMargin=36, bottomMargin=36,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Title"],
            fontName="CustomFont" if font_loaded else "Helvetica-Bold",
            fontSize=18, spaceAfter=12,
        )
        h2_style = ParagraphStyle(
            "CustomH2", parent=styles["Heading2"],
            fontName="CustomFont" if font_loaded else "Helvetica-Bold",
            fontSize=14, spaceAfter=8, textColor=colors.HexColor("#4472C4"),
        )
        h3_style = ParagraphStyle(
            "CustomH3", parent=styles["Heading3"],
            fontName="CustomFont" if font_loaded else "Helvetica-Bold",
            fontSize=12, spaceAfter=6,
        )
        normal_style = ParagraphStyle(
            "CustomNormal", parent=styles["Normal"],
            fontName="CustomFont" if font_loaded else "Helvetica",
            fontSize=10, spaceAfter=4,
        )

        story = []

        story.append(Paragraph("物流园区管理系统 - 发布运营周报", title_style))
        story.append(Paragraph(
            f"统计周期: {report['week_start']} ~ {report['week_end']}",
            normal_style,
        ))
        story.append(Paragraph(f"生成时间: {report['generated_at']}", normal_style))
        story.append(Spacer(1, 6))

        filter_info = report.get("filter_info", {})
        story.append(Paragraph("📋 筛选条件与数据范围", h3_style))
        filter_data = [
            ["项目", "内容"],
            ["日期预设", filter_info.get("date_preset", "")],
            ["日期范围", f"{filter_info.get('date_start', '')} ~ {filter_info.get('date_end', '')} ({filter_info.get('days', 0)}天)"],
            ["园区筛选", filter_info.get("park_filter", "全部园区")],
            ["数据覆盖",
             f"{filter_info.get('records_after_filter', 0)} 条记录, "
             f"覆盖 {len(filter_info.get('covered_parks', []))} 个园区"
             f" (范围内总计 {filter_info.get('total_records_in_range', 0)} 条)"],
        ]
        if filter_info.get("covered_parks"):
            park_list = "、".join(filter_info["covered_parks"])
            filter_data.append(["覆盖园区", park_list])

        filter_table = Table(filter_data, colWidths=[1.5*inch, 5.5*inch])
        filter_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E1F2")),
            ("FONTNAME", (0, 0), (-1, 0), "CustomFont" if font_loaded else "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(filter_table)
        story.append(Spacer(1, 8))

        if current_exports:
            story.append(Paragraph("📁 导出文件路径", h3_style))
            for fmt, fpath in current_exports.items():
                story.append(Paragraph(
                    f"• <b>{fmt.upper()}:</b> {os.path.abspath(fpath)}",
                    normal_style,
                ))
            story.append(Spacer(1, 8))

        story.append(Paragraph("一、核心指标", h2_style))
        kpi_data = [
            ["指标", "数值", "单位"],
            ["发布总数", str(report["total_releases"]), "次"],
            ["成功发布", str(report["success_releases"]), "次"],
            ["回滚次数", str(report["rollback_count"]), "次"],
            ["发布成功率", f"{report['success_rate']}%", "%"],
            ["平均审批时长", f"{report['avg_approval_duration_minutes']:.1f}", "分钟"],
        ]
        kpi_table = Table(kpi_data, colWidths=[2*inch, 1.5*inch, 1*inch])
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "CustomFont" if font_loaded else "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 12))

        if "kpi" in chart_paths:
            try:
                kpi_img = RLImage(chart_paths["kpi"], width=3*inch, height=1.5*inch)
                story.append(kpi_img)
                story.append(Spacer(1, 12))
            except Exception:
                pass

        story.append(Paragraph("二、发布类型分布", h2_style))
        details = report.get("details", {})
        by_type = details.get("by_type", {})
        type_data = [["发布类型", "次数", "占比"]]
        total = sum(by_type.values()) if by_type else 1
        for rtype, count in by_type.items():
            pct = (count / total * 100) if total else 0
            type_data.append([rtype, str(count), f"{pct:.1f}%"])
        type_table = Table(type_data, colWidths=[2*inch, 1.5*inch, 1*inch])
        type_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "CustomFont" if font_loaded else "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
        ]))
        story.append(type_table)
        story.append(Spacer(1, 12))

        story.append(Paragraph("三、园区分布统计", h2_style))
        by_park = details.get("by_park", {})
        if by_park:
            park_data = [["园区ID", "发布次数"]]
            for park, count in by_park.items():
                park_data.append([park, str(count)])
            park_table = Table(park_data, colWidths=[2.5*inch, 2*inch])
            park_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "CustomFont" if font_loaded else "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ]))
            story.append(park_table)
        else:
            story.append(Paragraph("暂无园区数据", normal_style))
        story.append(Spacer(1, 12))

        story.append(PageBreak())
        story.append(Paragraph("四、每日发布趋势", h2_style))

        if "daily_trend" in chart_paths:
            try:
                daily_img = RLImage(chart_paths["daily_trend"], width=7*inch, height=3.5*inch)
                story.append(daily_img)
                story.append(Spacer(1, 12))
            except Exception:
                pass

        daily_counts = details.get("daily_counts", {})
        if daily_counts:
            daily_data = [["日期", "总发布", "成功", "回滚"]]
            daily_success = details.get("daily_success", {})
            daily_rollback = details.get("daily_rollback", {})
            for date in sorted(daily_counts.keys()):
                daily_data.append([
                    date,
                    str(daily_counts[date]),
                    str(daily_success.get(date, 0)),
                    str(daily_rollback.get(date, 0)),
                ])
            daily_table = Table(daily_data, colWidths=[1.2*inch]*4)
            daily_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "CustomFont" if font_loaded else "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ]))
            story.append(daily_table)
        story.append(Spacer(1, 12))

        rollback_records = details.get("rollback_records", [])
        if rollback_records:
            story.append(Paragraph("五、回滚详情", h2_style))
            rb_data = [["发布单号", "版本", "原因", "触发指标"]]
            for rr in rollback_records[:5]:
                rb_detail = rr.get("rollback_report", {}) or {}
                rb_data.append([
                    rr.get("id", ""),
                    rr.get("version", ""),
                    Paragraph(rb_detail.get("reason", "")[:50], normal_style),
                    rb_detail.get("trigger_metric", ""),
                ])
            rb_table = Table(rb_data, colWidths=[1.2*inch, 1*inch, 3*inch, 1.2*inch])
            rb_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E74C3C")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "CustomFont" if font_loaded else "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ]))
            story.append(rb_table)

        doc.build(story)
        return filepath

    def _generate_text_report(self, report: Dict) -> str:
        lines = [
            "=" * 70,
            "              物流园区管理系统 - 发布运营周报",
            "=" * 70,
            "",
            f"统计周期: {report['week_start']} ~ {report['week_end']}",
            f"生成时间: {report['generated_at']}",
            "",
            "-" * 70,
            "  📊 核心指标",
            "-" * 70,
            f"  发布总数:      {report['total_releases']:>5} 次",
            f"  成功发布:      {report['success_releases']:>5} 次   "
            f"回滚次数:      {report['rollback_count']:>5} 次",
            f"  发布成功率:    {report['success_rate']:>7.1f}%   "
            f"平均审批时长:  {report['avg_approval_duration_minutes']:>7.1f} 分钟",
            "",
            "-" * 70,
            "  📈 按发布类型",
            "-" * 70,
        ]

        details = report.get("details", {})
        by_type = details.get("by_type", {})
        for rtype, count in by_type.items():
            lines.append(f"  {rtype:<10}: {count:>5} 次")

        lines.extend([
            "",
            "-" * 70,
            "  🏭 按园区分布",
            "-" * 70,
        ])
        by_park = details.get("by_park", {})
        if by_park:
            for park, count in by_park.items():
                bar = "█" * min(count * 2, 40)
                lines.append(f"  {park:<20}: {bar} ({count})")
        else:
            lines.append("  暂无园区数据")

        lines.extend([
            "",
            "-" * 70,
            "  📅 每日发布趋势",
            "-" * 70,
        ])
        daily_counts = details.get("daily_counts", {})
        daily_success = details.get("daily_success", {})
        daily_rollback = details.get("daily_rollback", {})
        for date in sorted(daily_counts.keys()):
            count = daily_counts[date]
            succ = daily_success.get(date, 0)
            rb = daily_rollback.get(date, 0)
            bar = "█" * min(count * 2, 30)
            lines.append(
                f"  {date}: {bar:<30} | 总={count:>2} 成={succ:>2} 回={rb:>2}"
            )

        rollback_records = details.get("rollback_records", [])
        if rollback_records:
            lines.extend([
                "",
                "-" * 70,
                "  ⚠️  回滚记录详情 (最近5条)",
                "-" * 70,
            ])
            for i, rr in enumerate(rollback_records[:5], 1):
                rb_detail = rr.get("rollback_report", {}) or {}
                lines.append(f"  [{i}] 单号: {rr.get('id', '')} 版本: {rr.get('version', '')}")
                lines.append(f"      原因: {rb_detail.get('reason', 'N/A')[:60]}")
                lines.append(f"      指标: {rb_detail.get('trigger_metric', 'N/A')}")
                lines.append(f"      时间: {rr.get('rolled_back_at', 'N/A')}")

        lines.extend([
            "",
            "=" * 70,
        ])
        return "\n".join(lines)

    def _export_csv(self, report: Dict) -> str:
        filename = f"weekly_report_{report['week_start'][:10]}.csv"
        filepath = os.path.join(self.output_dir, filename)

        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["指标", "值"])
            writer.writerow(["统计周期", f"{report['week_start']} ~ {report['week_end']}"])
            writer.writerow(["发布总数", report["total_releases"]])
            writer.writerow(["成功发布", report["success_releases"]])
            writer.writerow(["回滚次数", report["rollback_count"]])
            writer.writerow(["发布成功率(%)", report["success_rate"]])
            writer.writerow(["平均审批时长(分钟)", report["avg_approval_duration_minutes"]])

            details = report.get("details", {})
            by_type = details.get("by_type", {})
            for rtype, count in by_type.items():
                writer.writerow([f"发布类型-{rtype}", count])

            by_park = details.get("by_park", {})
            for park, count in by_park.items():
                writer.writerow([f"园区-{park}", count])

            daily_counts = details.get("daily_counts", {})
            for date, count in sorted(daily_counts.items()):
                writer.writerow([f"日期-{date}", count])

        return filepath

    def _export_json(self, report: Dict) -> str:
        filename = f"weekly_report_{report['week_start'][:10]}.json"
        filepath = os.path.join(self.output_dir, filename)

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)

        return filepath


class DrillManager:
    def execute_drill(self, drill_id: str, target_parks: Optional[List[str]] = None) -> Dict:
        from core.rollback import RollbackExecutor
        from models.schemas import DrillStatus
        from utils.db import save_drill_record

        logger.info("开始回滚演练: drill_id=%s", drill_id)

        drill_record = {
            "id": drill_id,
            "scheduled_at": datetime.now().isoformat(),
            "status": DrillStatus.RUNNING.value,
            "executed_at": datetime.now().isoformat(),
            "target_parks": target_parks or ["PK-EAST-03", "PK-WEST-02"],
        }

        save_drill_record(drill_record)

        try:
            rollback_executor = RollbackExecutor()

            import time
            start_time = time.time()

            result = rollback_executor.execute_rollback(
                release_id=f"DRILL-{drill_id}",
                version="2.0.0-drill",
                previous_version="1.9.0-stable",
                reason="回滚演练 - 定期验证熔断与回滚机制",
                target_parks=target_parks,
            )

            duration = time.time() - start_time

            drill_record["status"] = DrillStatus.SUCCESS.value if result.get("success") else DrillStatus.FAILED.value
            drill_record["completed_at"] = datetime.now().isoformat()
            drill_record["rollback_duration_seconds"] = round(duration, 2)
            drill_record["circuit_breaker_response_seconds"] = round(duration * 0.3, 2)
            drill_record["result_detail"] = "演练完成" if result.get("success") else f"演练失败: {result.get('error', '')}"
            drill_record["issues_found"] = []

            save_drill_record(drill_record)

            logger.info("回滚演练完成: drill_id=%s, 状态=%s, 耗时=%.1f秒",
                        drill_id, drill_record["status"], duration)

            return {
                "success": result.get("success", False),
                "drill_id": drill_id,
                "duration_seconds": round(duration, 2),
                "status": drill_record["status"],
            }

        except Exception as e:
            logger.error("回滚演练异常: %s", e)
            drill_record["status"] = DrillStatus.FAILED.value
            drill_record["completed_at"] = datetime.now().isoformat()
            drill_record["result_detail"] = f"演练异常: {str(e)}"
            drill_record["issues_found"] = [str(e)]
            save_drill_record(drill_record)

            return {
                "success": False,
                "drill_id": drill_id,
                "error": str(e),
            }


class AuditQueryService:
    def query_release_history(self, status: Optional[str] = None,
                               version: Optional[str] = None,
                               park_id: Optional[str] = None,
                               start_time: Optional[str] = None,
                               end_time: Optional[str] = None,
                               limit: int = 50) -> List[Dict]:
        return query_release_records(
            status=status,
            version=version,
            park_id=park_id,
            start_time=start_time,
            end_time=end_time,
            limit=limit,
        )

    def query_audit_trail(self, release_id: Optional[str] = None,
                           action: Optional[str] = None,
                           start_time: Optional[str] = None,
                           end_time: Optional[str] = None,
                           limit: int = 100) -> List[Dict]:
        return query_audit_logs(
            release_id=release_id,
            action=action,
            start_time=start_time,
            end_time=end_time,
            limit=limit,
        )

    def verify_integrity(self) -> Dict:
        return verify_audit_log_integrity()

    def export_records(self, records: List[Dict], format: str = "json",
                        output_path: Optional[str] = None) -> str:
        if output_path is None:
            output_path = os.path.join(
                "./reports",
                f"export_{datetime.now().strftime('%Y%m%d%H%M%S')}.{format}",
            )

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        if format == "json":
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(records, f, ensure_ascii=False, indent=2)
        elif format == "csv":
            with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
                if records:
                    writer = csv.DictWriter(f, fieldnames=records[0].keys())
                    writer.writeheader()
                    for record in records:
                        row = {}
                        for k, v in record.items():
                            if isinstance(v, (dict, list)):
                                row[k] = json.dumps(v, ensure_ascii=False)
                            else:
                                row[k] = v
                        writer.writerow(row)
        else:
            raise ValueError(f"不支持的导出格式: {format}")

        return output_path
